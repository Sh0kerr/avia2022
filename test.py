from openpyxl import load_workbook
import numpy as np

wb_flights = load_workbook('./flights.xlsx')
wb_points_and_distances = load_workbook('./distance.xlsx')

sheet_flights = wb_flights[wb_flights.sheetnames[0]]
sheet_points = wb_points_and_distances[wb_points_and_distances.sheetnames[0]]
sheet_ranges = wb_points_and_distances[wb_points_and_distances.sheetnames[1]]

for flight in range(2, sheet_flights.max_row + 1):
	plan_time = sheet_flights.cell(row=flight, column=6).value
	task_type = sheet_flights.cell(row=flight, column=2).value

	if task_type == 'A':
		location_dispatch_location_id = sheet_flights.cell(row=flight, column=10).value
		location_destination_location_id = sheet_flights.cell(row=flight, column=11).value
	else:
		location_dispatch_location_id = sheet_flights.cell(row=flight, column=11).value
		location_destination_location_id = sheet_flights.cell(row=flight, column=10).value

	for location in range(2, sheet_points.max_row + 1):
		if sheet_points.cell(row=location, column=2).value == location_dispatch_location_id:
			location_dispatch_point_id = sheet_points.cell(row=location, column=1).value
		elif sheet_points.cell(row=location, column=2).value == location_destination_location_id:
			location_destination_point_id = sheet_points.cell(row=location, column=1).value

	ranges_amount = sheet_ranges.max_row + 1
	arr = np.zeros((ranges_amount, ranges_amount))
	points = set()
	for distance in range(2, ranges_amount):
		points.add(sheet_ranges.cell(row=distance, column=2).value)
		if (sheet_ranges.cell(row=distance, column=2).value == location_destination_point_id) and (sheet_ranges.cell(row=distance, column=3).value == location_dispatch_point_id):
			hours = plan_time.hour
			minutes = plan_time.minute

points = sorted(points)
