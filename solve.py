from typing import Tuple
from task import TaskInfo
from openpyxl import load_workbook


class TaskGenerator:
	"""
	plan_time - плановое время (время прилета самолета)
	task_type - определяем взлет или посадка (важно при расчете расстояния, т.к. они по направлению разные)
	location_dispatch_location_id - имя локации начала маршрута
	location_destination_location_id - имя локации конца маршрута
	location_dispatch_point_id - номер точки начала маршрута
	location_destination_point_id - номер точки конца маршрута
	"""
	def get_data(self) -> Tuple[TaskInfo]:
		wb_flights = load_workbook('./Flights.xlsx')
		wb_points_and_distances = load_workbook('./Distance.xlsx')

		sheet_flights = wb_flights[wb_flights.sheetnames[0]]
		sheet_points = wb_points_and_distances[wb_points_and_distances.sheetnames[0]]

		for flight in range(2, sheet_flights.max_row):
			# plan_time = sheet_flights.cell(row=flight, column=6).value - 15 - время_на_доехать - 5
			task_type = sheet_flights.cell(row=flight, column=2).value

			if task_type == 'A':
				location_dispatch_location_id = sheet_flights.cell(row=flight, column=10).value
				location_destination_location_id = sheet_flights.cell(row=flight, column=11).value
			else:
				location_dispatch_location_id = sheet_flights.cell(row=flight, column=11).value
				location_destination_location_id = sheet_flights.cell(row=flight, column=10).value

			for location in range(2, sheet_points.max_row):
				if sheet_points.cell(row=location, column=2).value == location_dispatch_location_id:
					location_dispatch_point_id = sheet_points.cell(row=location, column=1).value
				elif sheet_points.cell(row=location, column=2).value == location_destination_location_id:
					location_destination_point_id = sheet_points.cell(row=location, column=1).value

		return result
